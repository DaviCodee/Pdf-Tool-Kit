"""Tabelas detectadas em PDF -> planilha .xlsx via openpyxl (extra ``xlsx``).

Recebe a saída de ``engines.tables.extract_tables`` (lista de ``(nome, csv)``) e
empacota em uma única planilha, com uma aba por tabela.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from pdftoolkit.core.errors import MissingDependencyError, OperationError


def _openpyxl() -> Any:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise MissingDependencyError(
            "pdf-to-xlsx requer o extra 'xlsx' (pip install pdftoolkit[xlsx])"
        ) from exc
    return Workbook


def tables_to_xlsx(tables: list[tuple[str, str]]) -> bytes:
    """Empacota ``tables`` (``[(nome_aba, csv_text), ...]``) em um único .xlsx.

    Nomes de abas são truncados a 31 caracteres (limite do Excel) e deduplicados.
    """
    Workbook = _openpyxl()
    workbook = Workbook()
    workbook.remove(workbook.active)

    used_names: set[str] = set()
    for original_name, csv_text in tables:
        sheet = workbook.create_sheet(title=_unique_sheet_name(original_name, used_names))
        for row in csv.reader(io.StringIO(csv_text)):
            sheet.append(["" if cell is None else cell for cell in row])

    if not workbook.sheetnames:
        # Nenhuma tabela: cria uma aba vazia para o arquivo não ficar inválido.
        workbook.create_sheet(title="vazio")

    buffer = io.BytesIO()
    try:
        workbook.save(buffer)
    except Exception as exc:
        raise OperationError(f"falha ao gerar xlsx: {exc}") from exc
    return buffer.getvalue()


def _unique_sheet_name(name: str, used: set[str]) -> str:
    """Sanitiza e garante unicidade entre os nomes de aba do workbook."""
    stem = name.rsplit(".", 1)[0]
    sanitized = "".join(ch if ch.isalnum() or ch in (" ", "_", "-") else "_" for ch in stem)
    sanitized = sanitized.strip()[:31] or "aba"
    candidate = sanitized
    suffix = 1
    while candidate in used:
        suffix += 1
        trimmed = sanitized[: max(1, 31 - len(f"-{suffix}"))]
        candidate = f"{trimmed}-{suffix}"
    used.add(candidate)
    return candidate
